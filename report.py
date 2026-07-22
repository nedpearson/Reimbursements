import json, os, datetime as _dt
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def build_workbook(rows, OUT):

    # ---- classify ----
    CREDIT_VENDOR='Paid TO Lindsey'
    credits=[r for r in rows if r['vendor']==CREDIT_VENDOR]
    billable=[r for r in rows if r['include'] and r.get('in_window',True) and r['amount'] is not None and r['vendor']!=CREDIT_VENDOR and r.get('pct') is not None]
    review=[r for r in rows if r['vendor']!=CREDIT_VENDOR and not (r['include'] and r.get('in_window',True) and r['amount'] is not None and r.get('pct') is not None)]
    for r in review:
        if r['include'] and not r.get('in_window',True) and not r.get('note'):
            r=r  # (note added below)
        if r.get('include') and not r.get('in_window',True):
            r['note']=(r.get('note') or '')+(' | ' if r.get('note') else '')+'before Aug 2024 cutoff'

    import json as _json, os as _os
    _cfg=_json.load(open(_os.path.join(_os.path.dirname(_os.path.abspath(__file__)),'config.json')))
    _sp=_cfg.get('split_percent',{})
    CATS=list(_sp.keys())
    DEFPCT={k:(v/100.0 if v>1 else v) for k,v in _sp.items()}

    # ---- styles ----
    ARIAL=lambda **k: Font(name='Arial', **k)
    H1=ARIAL(bold=True,size=16,color='1F3864')
    H2=ARIAL(bold=True,size=12,color='1F3864')
    HDR=ARIAL(bold=True,color='FFFFFF')
    hdrfill=PatternFill('solid',fgColor='1F3864')
    inputfill=PatternFill('solid',fgColor='FFF2CC')
    bluef=ARIAL(color='0000FF',bold=True)
    totfill=PatternFill('solid',fgColor='D9E1F2')
    thin=Side(style='thin',color='BFBFBF')
    box=Border(left=thin,right=thin,top=thin,bottom=thin)
    MONEY='$#,##0.00;($#,##0.00);-'
    PCT='0%'
    def style_header(ws,row,ncol):
        for c in range(1,ncol+1):
            cell=ws.cell(row=row,column=c); cell.font=HDR; cell.fill=hdrfill; cell.border=box
            cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)

    wb=openpyxl.Workbook()

    # ============ SETTINGS ============
    ws=wb.active; ws.title='Settings'
    ws['A1']='Reimbursement Settings — edit the yellow cells; every sheet recalculates automatically'
    ws['A1'].font=H1
    ws['A3']='Percentage Lindsey reimburses, by category'; ws['A3'].font=H2
    ws['A4']='Category'; ws['B4']='% She Owes'
    style_header(ws,4,2)
    r=5; catrow={}
    for cat in CATS:
        ws.cell(r,1,cat).font=ARIAL()
        c=ws.cell(r,2,DEFPCT[cat]); c.font=bluef; c.fill=inputfill; c.number_format=PCT; c.border=box
        ws.cell(r,1).border=box
        catrow[cat]=r; r+=1
    SET_FIRST, SET_LAST = 5, r-1
    ws['A'+str(r+1)]='Subtract money you already paid Lindsey (Venmo credits)?'
    ws['A'+str(r+1)].font=ARIAL(bold=True)
    tog=ws.cell(r+1,2,('Yes' if _cfg.get('subtract_payments_to_lindsey') else 'No')); tog.font=bluef; tog.fill=inputfill; tog.border=box
    TOGGLE_CELL='Settings!$B$'+str(r+1)
    ws['A'+str(r+3)]='Notes: Household/utilities default 50%. School/tuition & medical default 12% (per your agreement).'
    ws['A'+str(r+3)].font=ARIAL(italic=True,size=9,color='808080')
    ws['A'+str(r+4)]='“Treat all bills as paid” is ON: tuition invoices are counted as paid; duplicate/partial receipts are not double-counted.'
    ws['A'+str(r+4)].font=ARIAL(italic=True,size=9,color='808080')
    ws.column_dimensions['A'].width=58; ws.column_dimensions['B'].width=14
    PCT_RANGE=f"Settings!$A${SET_FIRST}:$B${SET_LAST}"
    PCT_CATCOL=f"Settings!$A${SET_FIRST}:$A${SET_LAST}"
    PCT_VALCOL=f"Settings!$B${SET_FIRST}:$B${SET_LAST}"

    # ============ LEDGER ============
    wl=wb.create_sheet('Ledger')
    wl['A1']='Itemized Ledger — every bill, by date'; wl['A1'].font=H1
    hdrs=['Date','Vendor','Category','Description','Amount','Counted','% She Owes','Her Share','Source File','Notes']
    for i,h in enumerate(hdrs,1): wl.cell(3,i,h)
    style_header(wl,3,len(hdrs))
    billable.sort(key=lambda x:(x['date'] or '9999', x['category']))
    rr=4
    for r_ in billable:
        wl.cell(rr,1,r_['date']).font=ARIAL()
        wl.cell(rr,2,r_['vendor']).font=ARIAL()
        wl.cell(rr,3,r_['category']).font=ARIAL()
        wl.cell(rr,4,r_['desc']).font=ARIAL()
        a=wl.cell(rr,5,r_['amount']); a.font=ARIAL(); a.number_format=MONEY
        wl.cell(rr,6,'Yes').font=ARIAL()
        if r_.get('flat_share') is not None:
            p=wl.cell(rr,7,'flat $100'); p.font=ARIAL(size=9)
            h=wl.cell(rr,8,r_['her_share']); h.font=ARIAL(); h.number_format=MONEY
        else:
            p=wl.cell(rr,7,f'=IFERROR(INDEX({PCT_VALCOL},MATCH(C{rr},{PCT_CATCOL},0)),0)'); p.font=ARIAL(); p.number_format=PCT
            h=wl.cell(rr,8,f'=IF(F{rr}="Yes",ROUND(E{rr}*G{rr},2),0)'); h.font=ARIAL(); h.number_format=MONEY
        wl.cell(rr,9,r_['file']).font=ARIAL(size=8,color='808080')
        wl.cell(rr,10,r_['note']).font=ARIAL(size=8,color='808080')
        for c in range(1,11): wl.cell(rr,c).border=box
        rr+=1
    LED_FIRST,LED_LAST=4,rr-1
    # totals row
    wl.cell(rr,4,'TOTALS').font=ARIAL(bold=True)
    t5=wl.cell(rr,5,f'=SUM(E{LED_FIRST}:E{LED_LAST})'); t5.font=ARIAL(bold=True); t5.number_format=MONEY
    t8=wl.cell(rr,8,f'=SUM(H{LED_FIRST}:H{LED_LAST})'); t8.font=ARIAL(bold=True); t8.number_format=MONEY
    for c in range(1,11): wl.cell(rr,c).fill=totfill; wl.cell(rr,c).border=box
    LED_TOTAL_HER=f'Ledger!$H${rr}'
    widths=[12,26,16,40,13,9,10,13,34,40]
    for i,w in enumerate(widths,1): wl.column_dimensions[get_column_letter(i)].width=w
    wl.freeze_panes='A4'

    # ============ CREDITS ============
    wc=wb.create_sheet('Credits (Paid to Lindsey)')
    wc['A1']='Money you already paid Lindsey (Venmo) — subtracted from what she owes'; wc['A1'].font=H1
    for i,h in enumerate(['Date','Description','Amount','Source'],1): wc.cell(3,i,h)
    style_header(wc,3,4)
    credits.sort(key=lambda x:x['date'] or '')
    cr=4
    for r_ in credits:
        wc.cell(cr,1,r_['date']).font=ARIAL(); wc.cell(cr,2,r_['desc']).font=ARIAL()
        a=wc.cell(cr,3,r_['amount']); a.font=ARIAL(); a.number_format=MONEY
        wc.cell(cr,4,r_['file']).font=ARIAL(size=8,color='808080')
        for c in range(1,5): wc.cell(cr,c).border=box
        cr+=1
    CR_F,CR_L=4,cr-1
    wc.cell(cr,2,'TOTAL CREDITS').font=ARIAL(bold=True)
    if CR_L>=CR_F:
        tc=wc.cell(cr,3,f'=SUM(C{CR_F}:C{CR_L})')
    else:
        wc.cell(cr-0,1,'').font=ARIAL()
        wc.cell(4,2,'None — advances to Lindsey are claimed as a category in the Ledger.').font=ARIAL(italic=True,color='808080')
        tc=wc.cell(cr,3,0)
    tc.font=ARIAL(bold=True); tc.number_format=MONEY
    for c in range(1,5): wc.cell(cr,c).fill=totfill; wc.cell(cr,c).border=box
    CRED_TOTAL=f"'Credits (Paid to Lindsey)'!$C${cr}"
    for i,w in enumerate([12,44,13,34],1): wc.column_dimensions[get_column_letter(i)].width=w
    wc.freeze_panes='A4'

    # ============ BY MONTH ============
    wm=wb.create_sheet('By Month')
    wm['A1']='Her Share by Month'; wm['A1'].font=H1
    for i,h in enumerate(['Month','Billed Amount','Her Share'],1): wm.cell(3,i,h)
    style_header(wm,3,3)
    months=sorted({(r_['date'] or '')[:7] for r_ in billable if r_['date']})
    mr=4
    for m in months:
        wm.cell(mr,1,m).font=ARIAL()
        b=wm.cell(mr,2,f'=SUMPRODUCT((LEFT(Ledger!$A${LED_FIRST}:$A${LED_LAST},7)="{m}")*Ledger!$E${LED_FIRST}:$E${LED_LAST})')
        b.font=ARIAL(); b.number_format=MONEY
        h=wm.cell(mr,3,f'=SUMPRODUCT((LEFT(Ledger!$A${LED_FIRST}:$A${LED_LAST},7)="{m}")*Ledger!$H${LED_FIRST}:$H${LED_LAST})')
        h.font=ARIAL(); h.number_format=MONEY
        for c in range(1,4): wm.cell(mr,c).border=box
        mr+=1
    wm.cell(mr,1,'TOTAL').font=ARIAL(bold=True)
    for col in (2,3):
        t=wm.cell(mr,col,f'=SUM({get_column_letter(col)}4:{get_column_letter(col)}{mr-1})'); t.font=ARIAL(bold=True); t.number_format=MONEY; 
    for c in range(1,4): wm.cell(mr,c).fill=totfill; wm.cell(mr,c).border=box
    for i,w in enumerate([14,18,16],1): wm.column_dimensions[get_column_letter(i)].width=w

    # ============ SUMMARY ============
    wsum=wb.create_sheet('Summary'); wb.move_sheet('Summary', -(len(wb.sheetnames)-1))
    wsum['A1']='REIMBURSEMENT SUMMARY'; wsum['A1'].font=H1
    wsum['A2']=f'Gerald "Ned" Pearson — bills in "House Bills" folder. Prepared {_dt.date.today().strftime("%B %d, %Y")}.'
    wsum['A2'].font=ARIAL(italic=True,size=10,color='606060')
    for i,h in enumerate(['Category','Total Billed','% She Owes','She Owes'],1): wsum.cell(4,i,h)
    style_header(wsum,4,4)
    sr=5
    for cat in CATS:
        wsum.cell(sr,1,cat).font=ARIAL()
        b=wsum.cell(sr,2,f'=SUMIF(Ledger!$C${LED_FIRST}:$C${LED_LAST},A{sr},Ledger!$E${LED_FIRST}:$E${LED_LAST})'); b.font=ARIAL(); b.number_format=MONEY
        if cat=='AT&T Business':
            p=wsum.cell(sr,3,'flat $100/mo'); p.font=ARIAL(size=9)
        else:
            p=wsum.cell(sr,3,f'=IFERROR(INDEX({PCT_VALCOL},MATCH(A{sr},{PCT_CATCOL},0)),0)'); p.font=ARIAL(); p.number_format=PCT
        h=wsum.cell(sr,4,f'=SUMIF(Ledger!$C${LED_FIRST}:$C${LED_LAST},A{sr},Ledger!$H${LED_FIRST}:$H${LED_LAST})'); h.font=ARIAL(); h.number_format=MONEY
        for c in range(1,5): wsum.cell(sr,c).border=box
        sr+=1
    # subtotal
    wsum.cell(sr,1,'Subtotal (she owes)').font=ARIAL(bold=True)
    sub=wsum.cell(sr,4,f'=SUM(D5:D{sr-1})'); sub.font=ARIAL(bold=True); sub.number_format=MONEY
    for c in range(1,5): wsum.cell(sr,c).fill=totfill; wsum.cell(sr,c).border=box
    SUB_ROW=sr
    # credits
    wsum.cell(sr+1,1,'Less: money you already paid Lindsey').font=ARIAL()
    cr_=wsum.cell(sr+1,4,f'=-IF({TOGGLE_CELL}="Yes",{CRED_TOTAL},0)'); cr_.font=ARIAL(); cr_.number_format=MONEY
    for c in range(1,5): wsum.cell(sr+1,c).border=box
    # net
    wsum.cell(sr+2,1,'NET — LINDSEY OWES YOU').font=ARIAL(bold=True,size=13,color='FFFFFF')
    net=wsum.cell(sr+2,4,f'=D{SUB_ROW}+D{sr+1}'); net.font=ARIAL(bold=True,size=13,color='FFFFFF'); net.number_format=MONEY
    netfill=PatternFill('solid',fgColor='2E7D32')
    for c in range(1,5): wsum.cell(sr+2,c).fill=netfill; wsum.cell(sr+2,c).border=box
    for i,w in enumerate([34,16,12,16],1): wsum.column_dimensions[get_column_letter(i)].width=w
    wsum.cell(sr+4,1,'How this is calculated:').font=ARIAL(bold=True,size=9)
    notes=['• Each bill was read from the folder; utility bills use the current-period charge (past-due balances excluded so nothing double-counts).',
     '• Percentages are set on the Settings tab — change them there and every number updates.',
     '• School/tuition and medical default to 12%; household costs to 50%.',
     '• Tuition invoices are treated as paid; duplicate files and partial-payment receipts are listed on the Review tab, not counted twice.',
     '• Direct support to Eli (~$7,700) and child support (~$1,100/mo through 5/1/2026) are excluded from this claim and noted for the record. See the Review tab for other items intentionally excluded (duplicates, business items you flagged, unpaid balances).']
    for k,n in enumerate(notes):
        wsum.cell(sr+5+k,1,n).font=ARIAL(size=9,color='606060')

    # ============ REVIEW ============
    wr=wb.create_sheet('Review (Excluded)')
    wr['A1']='Items reviewed but NOT counted (duplicates, invoices already paid, unpaid balances)'; wr['A1'].font=H1
    wr['A2']='Shown for full transparency. Nothing here is added to the total.'; wr['A2'].font=ARIAL(italic=True,size=10,color='606060')
    for i,h in enumerate(['Date','Vendor','Category','Description','Amount','Why excluded','Source File'],1): wr.cell(4,i,h)
    style_header(wr,4,7)
    review.sort(key=lambda x:(x['category'],x['date'] or ''))
    xr=5
    for r_ in review:
        wr.cell(xr,1,r_['date']).font=ARIAL(); wr.cell(xr,2,r_['vendor']).font=ARIAL()
        wr.cell(xr,3,r_['category']).font=ARIAL(); wr.cell(xr,4,r_['desc']).font=ARIAL()
        a=wr.cell(xr,5,r_['amount']); a.font=ARIAL(); a.number_format=MONEY
        wr.cell(xr,6,r_['note']).font=ARIAL(size=9,color='606060'); wr.cell(xr,7,r_['file']).font=ARIAL(size=8,color='808080')
        for c in range(1,8): wr.cell(xr,c).border=box
        xr+=1
    for i,w in enumerate([12,26,16,40,13,40,34],1): wr.column_dimensions[get_column_letter(i)].width=w
    wr.freeze_panes='A5'

    # ============ MISSING BILLS ============
    wmiss=wb.create_sheet('Missing Bills')
    wmiss['A1']='Missing Bills — months with no bill on file (Aug 2024 – Jul 2026)'; wmiss['A1'].font=H1
    wmiss['A2']='Collect these and drop them in the folder, then re-run. "started" = first bill on file for that account.'
    wmiss['A2'].font=ARIAL(italic=True,size=10,color='606060')
    mos=[]; yy,mm=2024,8
    while (yy,mm)<=(2026,7):
        mos.append(f"{yy:04d}-{mm:02d}"); mm+=1
        if mm>12: mm=1; yy+=1
    def cov(pred):
        return {r['date'][:7] for r in rows if pred(r) and r.get('date')}
    checks=[
     ('Atmos (Gas)', lambda r:r['vendor']=='Atmos (Gas)'),
     ('Entergy (Electric)', lambda r:r['vendor']=='Entergy (Electric)'),
     ('AT&T Internet', lambda r:r['vendor']=='AT&T Internet'),
     ('BR Water — main', lambda r:r['vendor']=='BR Water' and 'Irrigation' not in (r.get('desc') or '')),
     ('BR Water — irrigation', lambda r:r['vendor']=='BR Water' and 'Irrigation' in (r.get('desc') or '')),
     ('Pool (Fernando)', lambda r:r['vendor']=='Pool (Fernando)'),
     ('PODS (Storage)', lambda r:str(r['vendor']).startswith('PODS')),
     ('AT&T Business', lambda r:str(r['vendor']).startswith('AT&T Business')),
    ]
    for i,h in enumerate(['Bill / account','On file','First on file','Missing months (need to collect)'],1): wmiss.cell(4,i,h)
    style_header(wmiss,4,4)
    rr=5
    for name,pred in checks:
        have=cov(pred); first=min(have) if have else '—'
        miss=[mo for mo in mos if mo not in have]
        wmiss.cell(rr,1,name).font=ARIAL()
        wmiss.cell(rr,2,f"{len(have)}/{len(mos)}").font=ARIAL()
        wmiss.cell(rr,3,first).font=ARIAL()
        c=wmiss.cell(rr,4,(", ".join(miss) if miss else "complete")); c.font=ARIAL(size=9)
        if miss: c.fill=PatternFill('solid',fgColor='FFF2CC')
        for cc in range(1,5): wmiss.cell(rr,cc).border=box; wmiss.cell(rr,cc).alignment=Alignment(vertical='top',wrap_text=True)
        rr+=1
    wmiss.cell(rr+1,1,'Note: bills dated the month AFTER service are normal; “missing” means no statement at all for that period.').font=ARIAL(italic=True,size=9,color='808080')
    for i,w in enumerate([26,10,14,95],1): wmiss.column_dimensions[get_column_letter(i)].width=w

    wb.save(OUT)
    return dict(billable=len(billable),credits=len(credits),review=len(review))
